#encoding=utf-8     #编码方式
from openpyxl import load_workbook  #导入模块
"""
    此类中getDataFromSheet方法，用于从Excel中读取测试数据。
"""

class ParseExcel(object):
    def __init__(self,excelPath,sheetName):
        #将要读取的Excel加载到内存
        self.wb = load_workbook(excelPath)
        #通过工作表名称获取一个工作表对象
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        #获取工作表中存在数据区域的最大行号
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        #用于存放从工作表中读出来的数据
        dataList = []
        #因为工作行的第一行是标题行，需要去掉
        for line in self.sheet.rows[1:]:
            #遍历工作表中数据区域的每一行，并将每行中各个单元格的数据取出存于列表tmpList中，
            #然后再将存放一行数据的列表添加到最终数据列表dataList中
            tmpList = []
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            dataList.append(tmpList)
        #将获取工作表中的所有数据的迭代对象返回
        return dataList

if __name__ == '__main__':
    pe = ParseExcel(u"测试数据.xlsx",u"登录用户")
    for i in pe.getDatasFromSheet():
        print i[0],i[1]